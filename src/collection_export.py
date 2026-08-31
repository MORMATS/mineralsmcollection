from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from src.item_types import item_type_label


EXPORT_HEADERS = (
    "ID interno",
    "Código",
    "Nombre de la pieza",
    "Tipo",
    "Mineral principal",
    "Estado",
    "Fecha de adquisición",
    "Origen de adquisición",
    "Precio de compra",
    "Precio de venta",
    "Fecha de venta",
    "Enlace de compra / anuncio",
    "Características especiales",
    "Minerales secundarios",
    "Notas de la pieza",
    "Localidad",
    "Mina / yacimiento",
    "Región",
    "País",
    "Latitud",
    "Longitud",
    "ID Mindat localidad",
    "Notas de la localidad",
    "URL de la localidad",
    "Fórmula",
    "Elementos",
    "Categoría",
    "Sistema cristalino",
    "Dureza mínima",
    "Dureza máxima",
    "Color",
    "Brillo",
    "Raya",
    "Chakras",
    "Signos zodiacales",
    "ID Mindat mineral",
    "ID RRUFF",
    "URL del mineral",
    "Número de fotos",
    "Fecha de alta",
)

TEXT_COLUMNS = {
    "Nombre de la pieza": 28,
    "Origen de adquisición": 24,
    "Enlace de compra / anuncio": 34,
    "Características especiales": 34,
    "Minerales secundarios": 28,
    "Notas de la pieza": 38,
    "Localidad": 24,
    "Mina / yacimiento": 26,
    "Región": 22,
    "Notas de la localidad": 36,
    "URL de la localidad": 34,
    "Fórmula": 22,
    "Elementos": 24,
    "Color": 24,
    "Chakras": 24,
    "Signos zodiacales": 24,
    "URL del mineral": 34,
}


def _safe_text(value: object | None) -> str:
    if value is None:
        return ""
    text = ILLEGAL_CHARACTERS_RE.sub("", str(value).strip())[:32767]
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _names(values: Iterable[object]) -> str:
    return ", ".join(
        name
        for value in values
        if (name := _safe_text(getattr(value, "name", "")))
    )


def collection_item_row(item: object) -> list[object]:
    mineral = item.mineral
    locality = item.locality
    return [
        item.id,
        _safe_text(item.item_code),
        _safe_text(item.display_name or mineral.name),
        item_type_label(item.item_type),
        _safe_text(mineral.name),
        "Vendido" if item.sold else "Disponible",
        item.acquisition_date,
        _safe_text(item.acquisition_source),
        item.purchase_price,
        item.sale_price,
        item.sold_at,
        _safe_text(item.purchase_link),
        _safe_text(item.special_features),
        _safe_text(item.secondary_minerals),
        _safe_text(item.notes),
        _safe_text(locality.name if locality else None),
        _safe_text(locality.mine if locality else None),
        _safe_text(locality.region if locality else None),
        _safe_text(locality.country if locality else None),
        locality.latitude if locality else None,
        locality.longitude if locality else None,
        locality.mindat_locality_id if locality else None,
        _safe_text(locality.notes if locality else None),
        _safe_text(locality.source_url if locality else None),
        _safe_text(mineral.formula),
        _safe_text(mineral.elements),
        _safe_text(mineral.category),
        _safe_text(mineral.crystal_system),
        mineral.hardness_min,
        mineral.hardness_max,
        _safe_text(mineral.color),
        _safe_text(mineral.luster),
        _safe_text(mineral.streak),
        _names(mineral.chakras),
        _names(mineral.zodiac_signs),
        mineral.mindat_id,
        _safe_text(mineral.rruff_id),
        _safe_text(mineral.source_url),
        len(item.images),
        item.created_at,
    ]


def build_collection_workbook(items: Iterable[object]) -> bytes:
    selected_items = list(items)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Piezas"
    sheet.sheet_view.showGridLines = False

    last_column = get_column_letter(len(EXPORT_HEADERS))
    sheet.merge_cells(f"A1:{last_column}1")
    title_cell = sheet["A1"]
    title_cell.value = "Colección de minerales · Exportación de piezas"
    title_cell.fill = PatternFill("solid", fgColor="173C35")
    title_cell.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34

    sheet.merge_cells(f"A2:{last_column}2")
    subtitle_cell = sheet["A2"]
    subtitle_cell.value = (
        f"{len(selected_items)} pieza(s) exportada(s) · "
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    subtitle_cell.fill = PatternFill("solid", fgColor="E9EFEB")
    subtitle_cell.font = Font(name="Aptos", size=10, color="405A50")
    subtitle_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 24

    header_row = 4
    for column_index, header in enumerate(EXPORT_HEADERS, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.fill = PatternFill("solid", fgColor="C8783E")
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 34

    thin_border = Border(bottom=Side(style="thin", color="D8E1DC"))
    for row_index, item in enumerate(selected_items, start=header_row + 1):
        for column_index, value in enumerate(collection_item_row(item), start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(name="Aptos", size=10, color="1D2925")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    last_row = header_row + max(len(selected_items), 1)
    if selected_items:
        table = Table(displayName="PiezasExportadas", ref=f"A{header_row}:{last_column}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        for column_index in range(1, len(EXPORT_HEADERS) + 1):
            sheet.cell(row=header_row + 1, column=column_index, value="")

    header_positions = {header: index + 1 for index, header in enumerate(EXPORT_HEADERS)}
    for header in ("Fecha de adquisición", "Fecha de venta"):
        column = header_positions[header]
        for row in range(header_row + 1, last_row + 1):
            sheet.cell(row=row, column=column).number_format = "dd/mm/yyyy"
    created_column = header_positions["Fecha de alta"]
    for row in range(header_row + 1, last_row + 1):
        sheet.cell(row=row, column=created_column).number_format = "dd/mm/yyyy hh:mm"

    for header in ("Precio de compra", "Precio de venta"):
        column = header_positions[header]
        for row in range(header_row + 1, last_row + 1):
            sheet.cell(row=row, column=column).number_format = '#,##0.00 [$€-es-ES]'

    for header in ("Latitud", "Longitud", "Dureza mínima", "Dureza máxima"):
        column = header_positions[header]
        for row in range(header_row + 1, last_row + 1):
            sheet.cell(row=row, column=column).number_format = "0.000000"

    for header in ("Enlace de compra / anuncio", "URL de la localidad", "URL del mineral"):
        column = header_positions[header]
        for row in range(header_row + 1, last_row + 1):
            cell = sheet.cell(row=row, column=column)
            if cell.value:
                cell.hyperlink = str(cell.value).lstrip("'")
                cell.style = "Hyperlink"

    status_column = header_positions["Estado"]
    for row in range(header_row + 1, last_row + 1):
        cell = sheet.cell(row=row, column=status_column)
        if cell.value == "Disponible":
            cell.fill = PatternFill("solid", fgColor="DCEBE3")
            cell.font = Font(name="Aptos", size=10, bold=True, color="17613C")
        elif cell.value == "Vendido":
            cell.fill = PatternFill("solid", fgColor="F3DDDA")
            cell.font = Font(name="Aptos", size=10, bold=True, color="8B2A24")

    for column_index, header in enumerate(EXPORT_HEADERS, start=1):
        default_width = max(11, min(len(header) + 3, 20))
        sheet.column_dimensions[get_column_letter(column_index)].width = TEXT_COLUMNS.get(
            header, default_width
        )

    sheet.freeze_panes = "C5"
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = f"1:{header_row}"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
