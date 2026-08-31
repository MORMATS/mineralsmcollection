from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from streamlit.testing.v1 import AppTest

from src import db as db_module
from src.collection_export import EXPORT_HEADERS, build_collection_workbook
from src.db import Base
from src.models import Chakra, CollectionItem, Locality, MineralSpecies, ZodiacSign


COLLECTION_PAGE_PATH = Path(__file__).resolve().parents[1] / "views" / "1_Coleccion.py"


def sample_export_item():
    mineral = SimpleNamespace(
        name="Cuarzo",
        formula="SiO₂",
        elements="Si, O",
        category="Silicatos",
        crystal_system="Trigonal",
        hardness_min=7.0,
        hardness_max=7.0,
        color="Incoloro",
        luster="Vítreo",
        streak="Blanca",
        chakras=[SimpleNamespace(name="Corona")],
        zodiac_signs=[SimpleNamespace(name="Leo")],
        mindat_id=3337,
        rruff_id="R040031",
        source_url="https://www.mindat.org/min-3337.html",
    )
    locality = SimpleNamespace(
        name="Almadén",
        mine="Mina Esperanza",
        region="Castilla-La Mancha",
        country="España",
        latitude=38.775,
        longitude=-4.831,
        mindat_locality_id=123,
        notes="Localidad histórica",
        source_url="https://www.mindat.org/loc-123.html",
    )
    return SimpleNamespace(
        id=24,
        item_code="MIN-0024",
        display_name="=Cuarzo maestro",
        item_type="mineral",
        mineral=mineral,
        locality=locality,
        acquisition_date=date(2025, 1, 2),
        acquisition_source="Feria",
        purchase_price=25.5,
        sale_price=None,
        sold=False,
        sold_at=None,
        purchase_link="https://example.com/pieza",
        special_features="Cristales definidos",
        secondary_minerals="Calcita",
        notes="Ejemplar\x00 de prueba",
        images=[SimpleNamespace(), SimpleNamespace()],
        created_at=datetime(2025, 1, 3, 10, 30),
    )


def test_collection_workbook_contains_formatted_filterable_piece_data():
    workbook_bytes = build_collection_workbook([sample_export_item()])

    assert workbook_bytes.startswith(b"PK")
    workbook = load_workbook(BytesIO(workbook_bytes))
    sheet = workbook["Piezas"]
    last_column = get_column_letter(len(EXPORT_HEADERS))

    assert sheet.freeze_panes == "C5"
    assert [sheet.cell(row=4, column=index).value for index in range(1, len(EXPORT_HEADERS) + 1)] == list(
        EXPORT_HEADERS
    )
    assert sheet["B5"].value == "MIN-0024"
    assert sheet["C5"].value == "'=Cuarzo maestro"
    assert sheet["F5"].value == "Disponible"
    assert sheet["O5"].value == "Ejemplar de prueba"
    assert sheet["G5"].value.date() == date(2025, 1, 2)
    assert "€" in sheet["I5"].number_format
    assert sheet["L5"].hyperlink.target == "https://example.com/pieza"
    assert sheet["AH5"].value == "Corona"
    assert sheet["AI5"].value == "Leo"
    assert sheet["AM5"].value == 2
    assert sheet.tables["PiezasExportadas"].ref == f"A4:{last_column}5"


def test_collection_excel_flow_exports_selected_filtered_items(monkeypatch):
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    test_session = sessionmaker(bind=engine, future=True)
    Base.metadata.create_all(engine)
    monkeypatch.setattr(db_module, "SessionLocal", test_session)

    with test_session() as db:
        mineral = MineralSpecies(
            name="Cuarzo",
            chakras=[Chakra(name="Corona")],
            zodiac_signs=[ZodiacSign(name="Leo")],
        )
        db.add(
            CollectionItem(
                item_code="MIN-0024",
                display_name="Cuarzo maestro",
                mineral=mineral,
                locality=Locality(country="España"),
                sold=False,
            )
        )
        db.commit()

    try:
        app = AppTest.from_file(COLLECTION_PAGE_PATH, default_timeout=20).run()
        assert not app.exception
        next(button for button in app.button if button.label == "Exportar Excel").click().run()

        assert not app.exception
        selector = next(widget for widget in app.multiselect if widget.label == "Piezas para exportar")
        assert selector.value == []
        selector.set_value(["MIN-0024"]).run()

        assert not app.exception
        assert any("El archivo incluirá 1 pieza" in caption.value for caption in app.caption)
    finally:
        engine.dispose()


def test_collection_shows_all_items_by_default(monkeypatch):
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    test_session = sessionmaker(bind=engine, future=True)
    Base.metadata.create_all(engine)
    monkeypatch.setattr(db_module, "SessionLocal", test_session)

    with test_session() as db:
        mineral = MineralSpecies(name="Amonite")
        db.add_all(
            [
                CollectionItem(
                    item_code=f"FOS-{index:04d}",
                    item_type="fossil",
                    mineral=mineral,
                    sold=False,
                )
                for index in range(1, 14)
            ]
        )
        db.commit()

    try:
        app = AppTest.from_file(COLLECTION_PAGE_PATH, default_timeout=20).run()

        assert not app.exception
        page_size = next(widget for widget in app.selectbox if widget.label == "Piezas por página")
        assert page_size.value == "Todos"
        assert any("13 resultado" in heading.value for heading in app.markdown)
    finally:
        engine.dispose()
